import requests
import openpyxl
import traceback
import constants


class Simblock:

    # Making PayTM API Request
    def make_paytm_request(self, mobile_number):
        try:
            response = requests.get(constants.PAYTM_URL, params={"channel": "web",
                                                                 "version": "2",
                                                                 "number": mobile_number,
                                                                 "child_state_id": "1",
                                                                 "site_id": "1",
                                                                 "locale": "en-in"})
            json_obj = response.json()
            if len(json_obj) == 0:
                return None
            else:
                return json_obj
        except Exception as e:
            print("Exception====>" + str(e))


    # Making Easemydeal API Request
    def make_easemydeal_request(self, mobile_number):
        try:
            response = requests.post(constants.EASEMYDEAL_URL, data={'mobile': mobile_number, 'service_type': 'mobile_prepaid'})
            json_obj = response.json()
            if len(json_obj) == 0:
                return None
            else:
                return json_obj
        except Exception as e:
            print("Exception====>" + str(e))


        # Making PayRup API Request
    def make_payrup_request(self, mobile_number):
        try:
            response = requests.post(constants.PAYRUP_URL, json={'phoneNumber': mobile_number})
            # print(response.content)
            json_obj = response.json()
            if json_obj['status'] is False:
                return None
            else:
                return json_obj
        except Exception as e:
            print("Exception====>" + str(e))


    # Opening XL File and Processing
    def process_xl_file(self, xl_file_path):

        try:
            excel_file = openpyxl.load_workbook(xl_file_path)
        except Exception as e:
            print("Exception====>" + str(e))
            exit(1)

        file_obj = excel_file.active
        sheet_obj = excel_file['Sheet1']
        for i in range(1, sheet_obj.max_row + 1):
            mobile_number_cell = sheet_obj.cell(row=i, column=1)
            LSA_cell = sheet_obj.cell(row=i, column=2)
            TSP_cell = sheet_obj.cell(row=i, column=3)

            if mobile_number_cell.value is None:
                continue

            if server == 1:
                json_obj = simblock_obj.make_paytm_request(mobile_number_cell.value)
                if json_obj is None:
                    continue
                LSA_cell.value = json_obj['Circle']
                TSP_cell.value = json_obj['Operator']

            elif server == 2:
                json_obj = simblock_obj.make_easemydeal_request(mobile_number_cell.value)
                if json_obj is None:
                    continue
                LSA_cell.value = json_obj['location']
                TSP_cell.value = json_obj['service']

            elif server == 3:
                json_obj = simblock_obj.make_payrup_request(mobile_number_cell.value)
                if json_obj is None:
                    continue
                LSA_cell.value = json_obj['result']['operator']
                TSP_cell.value = json_obj['result']['circle']

            print('Processing ===> ' + str(mobile_number_cell.value) + ' (' + str(i - 1) + ' of ' \
                  + str(simblock_obj.get_maximum_rows(sheet_obj) - 1) + ')  ' + \
                  str(round(i / simblock_obj.get_maximum_rows(sheet_obj) * 100)) + ' %' + ' Completed.')
        excel_file.save(constants.PROCESSED_XL_FILE_PATH)
        print('Completed')


    # Getting Number of Occupied Rows
    def get_maximum_rows(self, sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows


simblock_obj = Simblock()
print("Please Select the server.\n\t 1.Paytm \n\t 2.EasemyDeal \n\t 3.PayRup")
try:
    server = int(input("Choose (1 / 2 / 3): "))
except Exception as e:
    print("Choose between the options available..!")
    exit()
match server:
    case 1:
        simblock_obj.process_xl_file(constants.XL_FILE_PATH)
    case 2:
        simblock_obj.process_xl_file(constants.XL_FILE_PATH)
    case 3:
        simblock_obj.process_xl_file(constants.XL_FILE_PATH)
    case _:
        print("Choose between the options available..!")
